import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
from database import (
    save_to_db, load_from_db, get_db_info, get_last_import,
    get_db_size_kb, clear_db, db_exists, EXPECTED_TABLES
)

st.set_page_config(page_title="Osebni finančni dashboard", layout="wide", page_icon="💶")

# ========== BARVE & KONSTANTE ==========

C_BLUE    = '#1F4E78'
C_GREEN   = '#548235'
C_RED     = '#C00000'
C_AMBER   = '#FFC000'
C_TEAL    = '#0D7377'
C_PURPLE  = '#7030A0'
C_ORANGE  = '#E97132'

PALETTE   = [C_BLUE, C_GREEN, C_RED, C_TEAL, C_PURPLE, C_ORANGE, C_AMBER]

# Custom CSS
st.markdown("""
<style>
    /* ---- header ---- */
    .main-header {
        font-size: 2rem;
        font-weight: 700;
        color: #1F4E78;
        text-align: center;
        letter-spacing: -0.5px;
        margin-bottom: 0.25rem;
    }
    .main-subheader {
        text-align: center;
        color: #666;
        font-size: 0.9rem;
        margin-bottom: 1rem;
    }

    /* ---- sidebar DB status ---- */
    .db-status-ok {
        background: linear-gradient(135deg, #d4edda, #c3e6cb);
        border-left: 4px solid #548235;
        padding: 0.5rem 0.75rem;
        border-radius: 0.4rem;
        font-size: 0.85rem;
        color: #155724;
    }
    .db-status-empty {
        background: linear-gradient(135deg, #fff3cd, #ffeaa7);
        border-left: 4px solid #ffc107;
        padding: 0.5rem 0.75rem;
        border-radius: 0.4rem;
        font-size: 0.85rem;
        color: #856404;
    }
    .sheet-status {
        font-size: 0.78rem;
        padding: 0.3rem 0 0 0;
        line-height: 1.8;
        opacity: 0.9;
    }

    /* ---- tab section headers ---- */
    .section-header {
        font-size: 1.1rem;
        font-weight: 600;
        color: #1F4E78;
        border-bottom: 2px solid #1F4E78;
        padding-bottom: 0.25rem;
        margin: 1rem 0 0.75rem 0;
    }

    /* ---- alert boxes ---- */
    .alert-info {
        background: #e8f4fd;
        border-left: 4px solid #1F4E78;
        padding: 0.6rem 0.8rem;
        border-radius: 0.3rem;
        font-size: 0.88rem;
        color: #1F4E78;
    }
</style>
""", unsafe_allow_html=True)


# ========== CHART HELPER ==========

def chart_layout(fig, title='', height=480, legend_bottom=False, **kwargs):
    """Aplicira konsistenten vizualni stil na vse Plotly grafe."""
    layout = dict(
        title=dict(text=title, font=dict(size=15, color=C_BLUE), x=0, xanchor='left', pad=dict(l=4)),
        height=height,
        margin=dict(l=8, r=16, t=48, b=40),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family='Inter, Segoe UI, sans-serif', size=12),
        hoverlabel=dict(bgcolor='rgba(30,30,30,0.92)', font_color='white',
                        font_size=12, bordercolor='rgba(255,255,255,0.15)'),
        xaxis=dict(showgrid=False, zeroline=False,
                   tickfont=dict(size=11), title_font=dict(size=12, color='#555')),
        yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)', zeroline=False,
                   tickfont=dict(size=11), title_font=dict(size=12, color='#555')),
        legend=dict(
            orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0,
            font=dict(size=11), bgcolor='rgba(0,0,0,0)', borderwidth=0
        ) if not legend_bottom else dict(
            orientation='h', yanchor='top', y=-0.15, xanchor='center', x=0.5,
            font=dict(size=11)
        ),
    )
    layout.update(kwargs)
    fig.update_layout(**layout)
    return fig

# ========== SESSION STATE ==========

def init_session_state():
    defaults = {
        'df': None,           # transactions (pripravljeno)
        'sp_df': None,        # sp_transactions (pripravljeno)
        'budget_df': None,    # budget_plan
        'income_df': None,    # income_history
        'accounts_df': None,  # accounts
        'db_loaded': False,   # ali je baza že naložena
        'raw_data': {},       # surovi dict vseh sheetov
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

init_session_state()


# ========== EXCEL MULTI-SHEET IMPORT ==========

SHEET_NAMES = {
    'transactions': 'transactions',
    'sp_transactions': 'sp_transactions',
    'budget_plan': 'budget_plan',
    'income_history': 'income_history',
    'accounts': 'accounts',
}

def load_excel_multisheet(uploaded_file):
    """Prebere vseh 5 sheetov iz Excel datoteke."""
    try:
        xls = pd.ExcelFile(uploaded_file)
        available = xls.sheet_names
        data = {}
        missing = []

        for key, sheet_name in SHEET_NAMES.items():
            if sheet_name in available:
                data[key] = pd.read_excel(xls, sheet_name=sheet_name)
            else:
                data[key] = None
                missing.append(sheet_name)

        return data, missing
    except Exception as e:
        st.error(f"Napaka pri branju Excel: {str(e)}")
        return None, []


# ========== PRIPRAVA PODATKOV ==========

def prepare_transactions(df):
    """Pripravi transactions sheet za analizo."""
    df = df.copy()

    col_mappings = {
        'year_month': ['year_month', 'Year+month', 'YearMonth'],
        'mesec': ['Mesec', 'mesec', 'Month'],
        'leto': ['leto', 'Leto', 'Year'],
        'datum': ['Datum', 'Date'],
        'tip': ['Prihodek/Odhodek', 'Tip', 'Type'],
        'znesek': ['Znesek', 'Amount', 'Vrednost'],
        'kategorija': ['Kategorija', 'Category'],
        'podkategorija': ['Podkategorija', 'Subcategory'],
        'zadeva': ['Zadeva', 'Description'],
        'nujnost': ['Nujnost', 'Urgency'],
        'opomba': ['Opomba', 'Note'],
    }

    for key, possible_names in col_mappings.items():
        if key not in df.columns:
            for name in possible_names:
                if name in df.columns:
                    df[key] = df[name]
                    break

    # Pretvori leto in mesec v int (SQLite jih vrne kot float)
    for col in ['leto', 'mesec']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            df = df[df[col].notna()].copy()
            df[col] = df[col].astype(int)

    if 'leto' in df.columns and 'mesec' in df.columns:
        df['YearMonth_sort'] = (
            df['leto'].astype(str) + '-' +
            df['mesec'].astype(str).str.zfill(2)
        )

    if 'znesek' in df.columns:
        df = df[df['znesek'].notna()].copy()

    # Izključi S.P. kategorijo iz osebnih financ
    if 'kategorija' in df.columns:
        df = df[df['kategorija'].str.strip().str.lower() != 'sp'].copy()

    return df


def prepare_sp_transactions(df):
    """Pripravi sp_transactions sheet za analizo."""
    df = df.copy()

    col_mappings = {
        'year_month': ['year_month', 'Year+month'],
        'mesec': ['Mesec', 'mesec'],
        'leto': ['leto', 'Leto'],
        'datum': ['Datum'],
        'tip': ['Prihodek/Odhodek'],
        'znesek': ['Znesek', 'Amount'],
        'zadeva': ['Zadeva'],
        'plan_rubrika': ['plan rubrika', 'plan_rubrika'],
        'kategorija': ['Kategorija'],
        'opomba': ['Opomba'],
    }

    for key, possible_names in col_mappings.items():
        if key not in df.columns:
            for name in possible_names:
                if name in df.columns:
                    df[key] = df[name]
                    break

    # Pretvori leto in mesec v int (SQLite jih vrne kot float)
    for col in ['leto', 'mesec']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
            df = df[df[col].notna()].copy()
            df[col] = df[col].astype(int)

    if 'leto' in df.columns and 'mesec' in df.columns:
        df['YearMonth_sort'] = (
            df['leto'].astype(str) + '-' +
            df['mesec'].astype(str).str.zfill(2)
        )

    if 'znesek' in df.columns:
        df = df[df['znesek'].notna()].copy()

    return df


def apply_prepared_data(raw_data):
    """Pripravi vse sheete in shrani v session state."""
    if raw_data.get('transactions') is not None:
        st.session_state.df = prepare_transactions(raw_data['transactions'])
    if raw_data.get('sp_transactions') is not None:
        st.session_state.sp_df = prepare_sp_transactions(raw_data['sp_transactions'])
    if raw_data.get('budget_plan') is not None:
        st.session_state.budget_df = raw_data['budget_plan'].copy()
    if raw_data.get('income_history') is not None:
        st.session_state.income_df = raw_data['income_history'].copy()
    if raw_data.get('accounts') is not None:
        st.session_state.accounts_df = raw_data['accounts'].copy()
    st.session_state.raw_data = raw_data
    st.session_state.db_loaded = True


# ========== AUTO-LOAD IZ DB ==========

if not st.session_state.db_loaded and db_exists():
    raw = load_from_db()
    if raw:
        apply_prepared_data(raw)


# ========== GRAFI ==========

def fmt_eur(val):
    """Formatiraj vrednost kot EUR string za hover."""
    return f'{val:,.0f} €'


_MN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Maj',6:'Jun',
       7:'Jul',8:'Avg',9:'Sep',10:'Okt',11:'Nov',12:'Dec'}

def ym_to_label(ym_series):
    """Pretvori '2026-04' → 'Apr '26' za x-os grafov."""
    def _fmt(ym):
        try:
            y, m = ym.split('-')
            return f"{_MN[int(m)]} '{y[2:]}"
        except Exception:
            return ym
    return ym_series.map(_fmt)


def create_monthly_trend(df, selected_years, selected_months, view_type='trend'):
    """Ustvari mesečni trend graf."""
    filtered = df.copy()
    if selected_years:
        filtered = filtered[filtered['leto'].isin(selected_years)]
    if selected_months:
        filtered = filtered[filtered['mesec'].isin(selected_months)]

    if view_type == 'trend':
        p = (filtered[filtered['tip'].str.lower() == 'prihodek']
             .groupby('YearMonth_sort')['znesek'].sum())
        o = (filtered[filtered['tip'].str.lower() == 'odhodek']
             .groupby('YearMonth_sort')['znesek'].sum())
        merged = pd.DataFrame({'Prihodki': p, 'Odhodki': o}).fillna(0).sort_index()
        merged['Bilanca'] = merged['Prihodki'] - merged['Odhodki']
        merged['Label'] = ym_to_label(merged.index.to_series())

        avg_bilanca = merged['Bilanca'].mean()

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=merged['Label'], y=merged['Prihodki'], name='Prihodki',
            marker_color=C_GREEN, marker_opacity=0.85,
            hovertemplate='<b>%{x}</b><br>Prihodki: %{y:,.0f} €<extra></extra>'
        ))
        fig.add_trace(go.Bar(
            x=merged['Label'], y=-merged['Odhodki'], name='Odhodki',
            marker_color=C_RED, marker_opacity=0.85,
            hovertemplate='<b>%{x}</b><br>Odhodki: %{y:,.0f} €<extra></extra>'
        ))
        fig.add_trace(go.Scatter(
            x=merged['Label'], y=merged['Bilanca'], name='Bilanca',
            line=dict(color=C_BLUE, width=2.5),
            mode='lines+markers', marker=dict(size=5, color=C_BLUE),
            hovertemplate='<b>%{x}</b><br>Bilanca: %{y:,.0f} €<extra></extra>'
        ))
        fig.add_hline(
            y=avg_bilanca, line_dash='dot', line_color=C_TEAL, line_width=1.5,
            annotation_text=f'  povp. {avg_bilanca:,.0f} €',
            annotation_font=dict(size=10, color=C_TEAL),
            annotation_position='right'
        )
        chart_layout(fig, title='Mesečni prihodki in odhodki',
                     barmode='relative', hovermode='x unified',
                     xaxis=dict(type='category', showgrid=False,
                                tickfont=dict(size=11)),
                     yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                zeroline=True, zerolinecolor='rgba(128,128,128,0.4)',
                                tickformat=',.0f', ticksuffix=' €'))
    else:
        prihodki_sum = filtered[filtered['tip'].str.lower() == 'prihodek']['znesek'].sum()
        odhodki_sum  = filtered[filtered['tip'].str.lower() == 'odhodek']['znesek'].sum()
        bilanca = prihodki_sum - odhodki_sum

        fig = go.Figure(data=[go.Bar(
            x=['💰 Prihodki', '💸 Odhodki', '📊 Bilanca'],
            y=[prihodki_sum, odhodki_sum, bilanca],
            marker_color=[C_GREEN, C_RED, C_BLUE if bilanca >= 0 else C_RED],
            marker_opacity=0.88,
            text=[fmt_eur(prihodki_sum), fmt_eur(odhodki_sum), fmt_eur(bilanca)],
            textposition='outside', textfont=dict(size=12, color=C_BLUE)
        )])
        chart_layout(fig, title='Skupna vsota za izbrano obdobje',
                     yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                tickformat=',.0f', ticksuffix=' €'))
    return fig


def create_category_chart(df, selected_years, selected_months,
                          selected_categories, chart_type='pie', group_by_year=False):
    """Ustvari analizo po kategorijah."""
    filtered = df[df['tip'].str.lower() == 'odhodek'].copy()
    if selected_years:
        filtered = filtered[filtered['leto'].isin(selected_years)]
    if selected_months:
        filtered = filtered[filtered['mesec'].isin(selected_months)]
    if selected_categories:
        filtered = filtered[filtered['kategorija'].isin(selected_categories)]

    if 'kategorija' not in filtered.columns:
        return None

    if group_by_year and len(selected_years) > 1:
        cat_year_data = filtered.groupby(['kategorija', 'leto'])['znesek'].sum().reset_index()
        top_cats = (selected_categories if selected_categories
                    else filtered.groupby('kategorija')['znesek'].sum().nlargest(10).index.tolist())
        cat_year_data = cat_year_data[cat_year_data['kategorija'].isin(top_cats)]
        sorted_cats = (cat_year_data.groupby('kategorija')['znesek'].sum()
                       .sort_values(ascending=False).index.tolist())

        fig = go.Figure()
        years = sorted(cat_year_data['leto'].unique())
        colors = ['#4C72B0', '#DD8452', '#55A868', '#C44E52', '#8172B2', '#937860']

        for idx, year in enumerate(years):
            yd = cat_year_data[cat_year_data['leto'] == year]
            yd = yd.set_index('kategorija').reindex(sorted_cats).reset_index()
            yd['znesek'] = yd['znesek'].fillna(0)
            fig.add_trace(go.Bar(
                name=str(year), x=yd['kategorija'], y=yd['znesek'],
                marker_color=colors[idx % len(colors)], marker_opacity=0.87,
                text=yd['znesek'].apply(lambda x: f'{x:,.0f} €' if x > 0 else ''),
                textposition='outside', textfont=dict(size=10),
                hovertemplate='<b>%{x}</b> · ' + str(year) + '<br>%{y:,.0f} €<extra></extra>'
            ))

        chart_layout(fig, title='Primerjava kategorij po letih', height=560,
                     barmode='group', hovermode='x unified',
                     bargap=0.15, bargroupgap=0.08,
                     xaxis=dict(tickangle=-38, tickfont=dict(size=11)),
                     yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                tickformat=',.0f', ticksuffix=' €'))
    else:
        cat_data = (filtered.groupby('kategorija')['znesek'].sum()
                    .reset_index().sort_values('znesek', ascending=False).head(15))

        if chart_type == 'pie':
            fig = px.pie(
                cat_data, values='znesek', names='kategorija',
                color_discrete_sequence=px.colors.qualitative.Set2,
                hole=0.35
            )
            fig.update_traces(
                textinfo='percent+label',
                hovertemplate='<b>%{label}</b><br>%{value:,.0f} €  (%{percent})<extra></extra>'
            )
            chart_layout(fig, title='Odhodki po kategorijah', height=480)
        else:
            cat_data_s = cat_data.sort_values('znesek', ascending=True)
            fig = go.Figure(go.Bar(
                x=cat_data_s['znesek'], y=cat_data_s['kategorija'],
                orientation='h',
                marker_color=C_BLUE, marker_opacity=0.82,
                text=cat_data_s['znesek'].apply(fmt_eur),
                textposition='outside', textfont=dict(size=10),
                hovertemplate='<b>%{y}</b><br>%{x:,.0f} €<extra></extra>'
            ))
            chart_layout(fig, title='TOP 15 kategorij po znesku', height=500,
                         xaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                    tickformat=',.0f', ticksuffix=' €'),
                         yaxis=dict(tickfont=dict(size=11)))
    return fig


# ========== BUDGET TRACKING ==========

def create_budget_tab(df, budget_df):
    """Budget tracking dashboard: Plan vs Actual po kategorijah."""
    if budget_df is None or budget_df.empty:
        st.warning("Ni podatkov o planu. Najprej uvozi Excel z listom 'budget_plan'.")
        return

    # Najdi plan stolpce (Plan 25, Plan 26 ...)
    plan_cols = {}
    for col in budget_df.columns:
        if str(col).startswith('Plan '):
            try:
                year = 2000 + int(str(col).replace('Plan ', '').strip())
                plan_cols[year] = col
            except ValueError:
                pass

    if not plan_cols:
        st.warning("Ni najdenih plan stolpcev. Pričakovano: 'Plan 25', 'Plan 26' ...")
        return

    # Kontrole na vrhu taba
    ctrl1, ctrl2, ctrl3 = st.columns([1, 1, 2])

    with ctrl1:
        selected_budget_year = st.selectbox(
            "Leto plana", sorted(plan_cols.keys()),
            index=len(plan_cols) - 1
        )

    plan_col = plan_cols[selected_budget_year]

    with ctrl2:
        period_type = st.radio(
            "Obdobje",
            ["Celo leto", "Do danes (YTD)", "Izbrani meseci"]
        )

    # Določi mesece za filter in pro-rata faktor
    current_month = datetime.now().month
    current_year = datetime.now().year
    month_names_map = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Maj',6:'Jun',
                       7:'Jul',8:'Avg',9:'Sep',10:'Okt',11:'Nov',12:'Dec'}

    if period_type == "Celo leto":
        filter_months = list(range(1, 13))
        pro_rata = 1.0
        period_label = f"celo leto {selected_budget_year}"
    elif period_type == "Do danes (YTD)":
        if selected_budget_year == current_year:
            today = datetime.now()
            day_of_year = today.timetuple().tm_yday
            days_in_year = 366 if selected_budget_year % 4 == 0 else 365
            pro_rata = day_of_year / days_in_year
            filter_months = list(range(1, current_month + 1))
            period_label = f"jan–{today.strftime('%d.%m.')} ({day_of_year}/{days_in_year} dni)"
        else:
            filter_months = list(range(1, 13))
            pro_rata = 1.0
            period_label = f"celo leto {selected_budget_year}"
    else:
        with ctrl3:
            filter_months = st.multiselect(
                "Meseci", list(range(1, 13)),
                default=list(range(1, current_month + 1)),
                format_func=lambda x: f"{x} – {month_names_map[x]}"
            )
        pro_rata = len(filter_months) / 12 if filter_months else 1.0
        period_label = f"izbrani meseci {selected_budget_year}"

    if not filter_months:
        st.info("Izberi vsaj en mesec.")
        return

    # Dejanski odhodki iz transactions
    actual = df[(df['tip'].str.lower() == 'odhodek') & (df['leto'] == selected_budget_year)].copy()
    actual = actual[actual['mesec'].isin(filter_months)]
    actual_by_cat_raw = actual.groupby('kategorija')['znesek'].sum()
    # Case-insensitive lookup: "Investicije" ujame "investicije" v transakcijah
    actual_norm = {k.strip().lower(): v for k, v in actual_by_cat_raw.items()}

    # Pripravi budget tabelo
    budget_cols = ['Kategorija', plan_col]
    if 'Podkategorija' in budget_df.columns:
        budget_cols.append('Podkategorija')

    budget = budget_df[budget_cols].copy()
    budget = budget[budget[plan_col].notna() & (budget[plan_col] > 0)].copy()
    budget = budget.rename(columns={plan_col: 'Plan_letni'})
    budget['Plan_obdobje'] = (budget['Plan_letni'] * pro_rata).round(2)
    budget['Actual'] = budget['Kategorija'].str.strip().str.lower().map(actual_norm).fillna(0)
    budget['Ostane'] = budget['Plan_obdobje'] - budget['Actual']
    budget['Pct'] = (
        (budget['Actual'] / budget['Plan_obdobje'] * 100)
        .replace([float('inf'), float('nan')], 0)
        .clip(0, 999)
    )
    budget['Status'] = budget['Pct'].apply(
        lambda p: '🔴 Prekoračeno' if p > 100 else ('🟡 Opozorilo' if p > 80 else '🟢 OK')
    )

    # ---- SKUPNE METRIKE ----
    total_plan = budget['Plan_obdobje'].sum()
    total_actual = budget['Actual'].sum()
    total_ostane = total_plan - total_actual
    total_pct = (total_actual / total_plan * 100) if total_plan > 0 else 0

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("📋 Plan (obdobje)", f"{total_plan:,.0f} €",
                  help=f"Letni plan × {pro_rata:.2f} = {period_label}")
    with m2:
        st.metric("💸 Porabljeno", f"{total_actual:,.0f} €")
    with m3:
        st.metric("💰 Ostane", f"{total_ostane:,.0f} €",
                  delta="v redu" if total_ostane >= 0 else "PREKORAČENO",
                  delta_color="normal" if total_ostane >= 0 else "inverse")
    with m4:
        st.metric("📊 % porabe", f"{total_pct:.1f}%")

    # ---- ALARMI ----
    overbudget = budget[budget['Pct'] > 100]['Kategorija'].tolist()
    near_limit = budget[(budget['Pct'] > 80) & (budget['Pct'] <= 100)]['Kategorija'].tolist()
    zero_actual = budget[budget['Actual'] == 0]['Kategorija'].tolist()

    if overbudget:
        st.error(f"🔴 **Prekoračene kategorije ({len(overbudget)}):** {', '.join(overbudget)}")
    if near_limit:
        st.warning(f"🟡 **Blizu limita >80% ({len(near_limit)}):** {', '.join(near_limit)}")

    # Opozorilo za budget kategorije ki nimajo nobene transakcije
    if 'kategorija' in df.columns:
        trans_cats  = set(df[df['leto'] == selected_budget_year]['kategorija']
                          .dropna().str.strip().str.lower().unique())
        budget_cats_orig = budget['Kategorija'].dropna().unique()
        unmatched = [c for c in budget_cats_orig if c.strip().lower() not in trans_cats]
        if unmatched:
            with st.expander(f"ℹ️ {len(unmatched)} budget kategorij brez transakcij v {selected_budget_year}", expanded=False):
                st.caption("Te kategorije so v planu, ampak nimajo transakcij za izbrano leto/mesece.")
                for cat in sorted(unmatched):
                    st.markdown(f"- `{cat}`")

    st.divider()

    # ---- HORIZONTALNI GRAF: Plan vs Actual ----
    budget_sorted = budget.sort_values('Plan_obdobje', ascending=True)

    bar_colors = [
        C_RED if p > 100 else (C_AMBER if p > 80 else C_GREEN)
        for p in budget_sorted['Pct']
    ]

    fig = go.Figure()
    # Plan — ozadje
    fig.add_trace(go.Bar(
        y=budget_sorted['Kategorija'], x=budget_sorted['Plan_obdobje'],
        name='Plan', orientation='h',
        marker_color=f'rgba(31,78,120,0.15)',
        marker_line=dict(color=f'rgba(31,78,120,0.5)', width=1),
        hovertemplate='<b>%{y}</b><br>Plan: %{x:,.0f} €<extra></extra>',
    ))
    # Actual — spredaj
    fig.add_trace(go.Bar(
        y=budget_sorted['Kategorija'], x=budget_sorted['Actual'],
        name='Porabljeno', orientation='h',
        marker_color=bar_colors, marker_opacity=0.88,
        text=budget_sorted['Pct'].apply(lambda p: f'{p:.0f}%'),
        textposition='outside', textfont=dict(size=10),
        hovertemplate='<b>%{y}</b><br>Porabljeno: %{x:,.0f} €<extra></extra>',
    ))

    h = max(380, len(budget_sorted) * 34 + 100)
    chart_layout(fig, title=f'Plan vs Porabljeno — {period_label}', height=h,
                 barmode='overlay', hovermode='y unified',
                 margin=dict(l=8, r=70, t=50, b=30),
                 xaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                            tickformat=',.0f', ticksuffix=' €'),
                 yaxis=dict(tickfont=dict(size=11)))
    st.plotly_chart(fig, use_container_width=True)

    # ---- PODROBNA TABELA ----
    st.subheader("📊 Podrobna tabela")

    show_cols = ['Kategorija', 'Plan_letni', 'Plan_obdobje', 'Actual', 'Ostane', 'Pct', 'Status']
    if 'Podkategorija' in budget.columns:
        show_cols = show_cols[:1] + ['Podkategorija'] + show_cols[1:]

    display = budget[show_cols].copy()
    rename_map = {
        'Plan_letni': 'Plan letni €',
        'Plan_obdobje': f'Plan ({period_label}) €',
        'Actual': 'Porabljeno €',
        'Ostane': 'Ostane €',
        'Pct': '% porabe',
    }
    display = display.rename(columns=rename_map)

    def color_budget_row(row):
        pct = row['% porabe']
        if pct > 100:
            return ['color: #ff6b6b; font-weight: 600' if c == '% porabe' else '' for c in row.index]
        elif pct > 80:
            return ['color: #f0a500; font-weight: 600' if c == '% porabe' else '' for c in row.index]
        return [''] * len(row)

    fmt = {
        'Plan letni €': '{:,.0f} €',
        f'Plan ({period_label}) €': '{:,.0f} €',
        'Porabljeno €': '{:,.0f} €',
        'Ostane €': '{:,.0f} €',
        '% porabe': '{:.1f}%',
    }

    st.dataframe(
        display.style.apply(color_budget_row, axis=1).format(fmt),
        use_container_width=True,
        hide_index=True
    )


# ========== S.P. ANALIZA ==========

def create_sp_tab(sp_df):
    """S.P. dobičkonosnost analiza."""
    if sp_df is None or sp_df.empty:
        st.warning("Ni podatkov S.P. Najprej uvozi Excel z listom 'sp_transactions'.")
        return

    month_names_map = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Maj',6:'Jun',
                       7:'Jul',8:'Avg',9:'Sep',10:'Okt',11:'Nov',12:'Dec'}

    # ---- FILTRI ----
    f1, f2 = st.columns([1, 2])
    with f1:
        available_years = sorted(sp_df['leto'].unique().tolist())
        selected_years_sp = st.multiselect(
            "Leta", available_years, default=available_years,
            key="sp_years"
        )
    with f2:
        available_months_sp = sorted(sp_df['mesec'].unique().tolist())
        selected_months_sp = st.multiselect(
            "Meseci", available_months_sp,
            format_func=lambda x: f"{x} – {month_names_map.get(x,'')}",
            key="sp_months"
        )

    filtered_sp = sp_df.copy()
    if selected_years_sp:
        filtered_sp = filtered_sp[filtered_sp['leto'].isin(selected_years_sp)]
    if selected_months_sp:
        filtered_sp = filtered_sp[filtered_sp['mesec'].isin(selected_months_sp)]

    if filtered_sp.empty:
        st.info("Ni podatkov za izbrano obdobje.")
        return

    prihodki_sp = filtered_sp[filtered_sp['tip'].str.lower() == 'prihodek']['znesek'].sum()
    odhodki_sp  = filtered_sp[filtered_sp['tip'].str.lower() == 'odhodek']['znesek'].sum()
    dobicek     = prihodki_sp - odhodki_sp
    marza       = (dobicek / prihodki_sp * 100) if prihodki_sp > 0 else 0

    # ---- METRIKE ----
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("💰 Prihodki S.P.", f"{prihodki_sp:,.0f} €")
    with m2:
        st.metric("💸 Odhodki S.P.", f"{odhodki_sp:,.0f} €")
    with m3:
        st.metric("📈 Dobiček", f"{dobicek:,.0f} €",
                  delta="pozitiven" if dobicek >= 0 else "izguba",
                  delta_color="normal" if dobicek >= 0 else "inverse")
    with m4:
        st.metric("📊 Marža", f"{marza:.1f}%")

    st.divider()

    # ---- GRAFI ----
    gtab1, gtab2, gtab3 = st.tabs(["📈 Mesečni P&L", "🏷️ Stroški po kategorijah", "📋 Tabela"])

    # GRAF 1: Mesečni P&L
    with gtab1:
        prihodki_mes = (filtered_sp[filtered_sp['tip'].str.lower() == 'prihodek']
                        .groupby('YearMonth_sort')['znesek'].sum())
        odhodki_mes  = (filtered_sp[filtered_sp['tip'].str.lower() == 'odhodek']
                        .groupby('YearMonth_sort')['znesek'].sum())

        pl = pd.DataFrame({'Prihodki': prihodki_mes, 'Odhodki': odhodki_mes}).fillna(0).sort_index()
        pl['Dobiček'] = pl['Prihodki'] - pl['Odhodki']
        pl['Label'] = ym_to_label(pl.index.to_series())

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=pl['Label'], y=pl['Prihodki'], name='Prihodki',
            marker_color=C_GREEN, marker_opacity=0.85,
            hovertemplate='<b>%{x}</b><br>Prihodki: %{y:,.0f} €<extra></extra>'
        ))
        fig.add_trace(go.Bar(
            x=pl['Label'], y=-pl['Odhodki'], name='Odhodki',
            marker_color=C_RED, marker_opacity=0.85,
            hovertemplate='<b>%{x}</b><br>Odhodki: %{y:,.0f} €<extra></extra>'
        ))
        fig.add_trace(go.Scatter(
            x=pl['Label'], y=pl['Dobiček'], name='Dobiček',
            mode='lines+markers',
            line=dict(color=C_BLUE, width=2.5),
            marker=dict(size=6, color=C_BLUE),
            hovertemplate='<b>%{x}</b><br>Dobiček: %{y:,.0f} €<extra></extra>'
        ))
        chart_layout(fig, title='Mesečni P&L — S.P.',
                     barmode='relative', hovermode='x unified',
                     xaxis=dict(type='category', showgrid=False, tickfont=dict(size=11)),
                     yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                zeroline=True, zerolinecolor='rgba(128,128,128,0.4)',
                                tickformat=',.0f', ticksuffix=' €'))
        st.plotly_chart(fig, use_container_width=True)

        # Letna P&L tabela
        if len(selected_years_sp) > 0:
            st.subheader("📊 Letni povzetek")
            rows = []
            for year in sorted(selected_years_sp):
                yd = filtered_sp[filtered_sp['leto'] == year]
                yp = yd[yd['tip'].str.lower() == 'prihodek']['znesek'].sum()
                yo = yd[yd['tip'].str.lower() == 'odhodek']['znesek'].sum()
                yd_val = yp - yo
                ym = (yd_val / yp * 100) if yp > 0 else 0
                rows.append({'Leto': year, 'Prihodki €': yp, 'Odhodki €': yo,
                             'Dobiček €': yd_val, 'Marža %': ym})
            annual_df = pd.DataFrame(rows)

            def color_sp_row(row):
                n = len(row)
                if row['Dobiček €'] < 0:
                    return [''] * (n - 2) + ['color: #ff6b6b; font-weight: 600', '']
                return [''] * (n - 2) + ['color: #6abf69; font-weight: 600', '']

            st.dataframe(
                annual_df.style
                    .apply(color_sp_row, axis=1)
                    .format({'Prihodki €': '{:,.0f} €', 'Odhodki €': '{:,.0f} €',
                             'Dobiček €': '{:,.0f} €', 'Marža %': '{:.1f}%'}),
                use_container_width=True, hide_index=True
            )

    # GRAF 2: Stroški po kategorijah
    with gtab2:
        odhodki_df = filtered_sp[filtered_sp['tip'].str.lower() == 'odhodek'].copy()

        if 'kategorija' not in odhodki_df.columns or odhodki_df.empty:
            st.info("Ni podatkov o kategorijah odhodkov.")
        else:
            cat_data = (odhodki_df.groupby('kategorija')['znesek'].sum()
                        .reset_index().sort_values('znesek', ascending=False))

            kat_tip = st.radio("Tip grafa", ["Pie", "Bar"], horizontal=True, key="sp_chart")

            if kat_tip == "Pie":
                fig2 = px.pie(cat_data, values='znesek', names='kategorija',
                              color_discrete_sequence=px.colors.qualitative.Set2,
                              hole=0.35)
                fig2.update_traces(
                    textinfo='percent+label',
                    hovertemplate='<b>%{label}</b><br>%{value:,.0f} €  (%{percent})<extra></extra>'
                )
                chart_layout(fig2, title='Stroški S.P. po kategorijah', height=450)
            else:
                cat_s = cat_data.sort_values('znesek', ascending=True)
                fig2 = go.Figure(go.Bar(
                    x=cat_s['znesek'], y=cat_s['kategorija'],
                    orientation='h', marker_color=C_RED, marker_opacity=0.82,
                    text=cat_s['znesek'].apply(fmt_eur),
                    textposition='outside', textfont=dict(size=10),
                    hovertemplate='<b>%{y}</b><br>%{x:,.0f} €<extra></extra>'
                ))
                chart_layout(fig2, title='Stroški S.P. po kategorijah', height=450,
                             xaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                        tickformat=',.0f', ticksuffix=' €'))
            st.plotly_chart(fig2, use_container_width=True)

            # % delež vsake kategorije
            total_odh = cat_data['znesek'].sum()
            cat_data['Delež %'] = cat_data['znesek'] / total_odh * 100
            cat_data.columns = ['Kategorija', 'Znesek €', 'Delež %']
            st.dataframe(
                cat_data.style.format({'Znesek €': '{:,.0f} €', 'Delež %': '{:.1f}%'}),
                use_container_width=True, hide_index=True
            )

    # GRAF 3: Raw tabela
    with gtab3:
        st.subheader("📋 S.P. transakcije")
        disp_cols = ['YearMonth_sort', 'datum', 'tip', 'znesek', 'kategorija', 'zadeva', 'opomba']
        disp_cols = [c for c in disp_cols if c in filtered_sp.columns]

        def color_tip(row):
            try:
                val = str(row['tip']).lower() if 'tip' in row.index else ''
            except Exception:
                val = ''
            if val == 'prihodek':
                return ['background-color: #e8f5e9'] * len(row)
            return ['background-color: #ffebee'] * len(row)

        st.dataframe(
            filtered_sp[disp_cols].sort_values('YearMonth_sort', ascending=False)
                .style.apply(color_tip, axis=1),
            use_container_width=True, height=450
        )
        st.metric("Skupno transakcij", len(filtered_sp))


# ========== INCOME HISTORY ==========

def create_income_tab(income_df):
    """Plača / prihodki tracking iz income_history sheeta."""
    if income_df is None or income_df.empty:
        st.info("Še ni podatkov o prihodkih.")
        st.markdown("""
        **Kako dodaš podatke:**
        Odpri Excel → sheet `income_history` in dodaj stolpce:
        `year_month, Datum, Vir, Neto, Plačnik, Komentar`
        """)
        return

    inc = income_df.copy()

    # Normalizacija stolpcev
    col_map = {
        'year_month': ['year_month'],
        'datum':      ['Datum', 'datum'],
        'vir':        ['Vir', 'vir'],
        'neto':       ['Neto', 'neto', 'Znesek', 'Amount'],
        'placnik':    ['Plačnik', 'placnik', 'Plačnik'],
        'komentar':   ['Komentar', 'komentar'],
    }
    for key, names in col_map.items():
        if key not in inc.columns:
            for n in names:
                if n in inc.columns:
                    inc[key] = inc[n]
                    break

    if 'neto' not in inc.columns:
        st.error("Ni stolpca 'Neto' v income_history.")
        return

    inc['neto'] = pd.to_numeric(inc['neto'], errors='coerce').fillna(0)

    # Leto in mesec iz year_month ("2026-04") ali datum
    if 'year_month' in inc.columns:
        inc['leto']  = inc['year_month'].astype(str).str[:4].apply(
            lambda x: int(x) if x.isdigit() else None)
        inc['mesec'] = inc['year_month'].astype(str).str[5:7].apply(
            lambda x: int(x) if x.isdigit() else None)
    inc = inc[inc['leto'].notna() & inc['mesec'].notna()].copy()
    inc['leto']  = inc['leto'].astype(int)
    inc['mesec'] = inc['mesec'].astype(int)
    inc['Label'] = ym_to_label(inc['year_month'].astype(str))

    # ---- FILTRI ----
    f1, f2 = st.columns([1, 2])
    with f1:
        avail_years = sorted(inc['leto'].unique().tolist())
        cur_y = datetime.now().year
        def_years = [cur_y] if cur_y in avail_years else avail_years[-1:]
        sel_years = st.multiselect("Leta", avail_years, default=def_years, key="inc_years")
    with f2:
        avail_vir = sorted(inc['vir'].dropna().unique().tolist()) if 'vir' in inc.columns else []
        sel_vir = st.multiselect("Vir prihodka", avail_vir, key="inc_vir",
                                 help="Pusti prazno za vse")

    fi = inc.copy()
    if sel_years:
        fi = fi[fi['leto'].isin(sel_years)]
    if sel_vir:
        fi = fi[fi['vir'].isin(sel_vir)]

    if fi.empty:
        st.info("Ni podatkov za izbrano obdobje.")
        return

    # ---- METRIKE ----
    total_neto = fi['neto'].sum()
    avg_mes    = fi.groupby('year_month')['neto'].sum().mean()
    st_mes     = fi['year_month'].nunique()
    max_mes    = fi.groupby('year_month')['neto'].sum().max()

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("💰 Skupaj neto", f"{total_neto:,.0f} €")
    with m2:
        st.metric("📈 Povp./mesec", f"{avg_mes:,.0f} €")
    with m3:
        st.metric("📅 Mesecev", st_mes)
    with m4:
        st.metric("🏆 Najboljši mesec", f"{max_mes:,.0f} €")

    st.divider()

    itab1, itab2, itab3 = st.tabs(["📈 Trend", "🏷️ Po virih", "📋 Tabela"])

    # TAB 1: Mesečni trend
    with itab1:
        mes_data = fi.groupby(['year_month', 'Label'])['neto'].sum().reset_index()
        mes_data = mes_data.sort_values('year_month')

        # Povprečna referenčna črta
        avg_val = mes_data['neto'].mean()

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=mes_data['Label'], y=mes_data['neto'],
            name='Neto prihodek',
            marker_color=C_GREEN, marker_opacity=0.85,
            text=mes_data['neto'].apply(fmt_eur),
            textposition='outside', textfont=dict(size=10),
            hovertemplate='<b>%{x}</b><br>%{y:,.0f} €<extra></extra>'
        ))
        fig.add_hline(
            y=avg_val, line_dash='dot', line_color=C_TEAL, line_width=1.5,
            annotation_text=f'  povp. {avg_val:,.0f} €',
            annotation_font=dict(size=10, color=C_TEAL),
            annotation_position='right'
        )
        chart_layout(fig, title='Mesečni neto prihodek',
                     xaxis=dict(type='category', showgrid=False, tickfont=dict(size=11)),
                     yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                tickformat=',.0f', ticksuffix=' €'))
        st.plotly_chart(fig, use_container_width=True)

        # Letni primerjalni graf (če >1 leto)
        if len(sel_years) > 1:
            st.subheader("Letna primerjava")
            annual = fi.groupby('leto')['neto'].agg(['sum', 'mean', 'count']).reset_index()
            annual.columns = ['Leto', 'Skupaj €', 'Povp. mesec €', 'Mesecev']

            fig2 = go.Figure(go.Bar(
                x=annual['Leto'].astype(str), y=annual['Skupaj €'],
                marker_color=C_BLUE, marker_opacity=0.85,
                text=annual['Skupaj €'].apply(fmt_eur),
                textposition='outside', textfont=dict(size=11),
                hovertemplate='<b>%{x}</b><br>%{y:,.0f} €<extra></extra>'
            ))
            chart_layout(fig2, title='Letni neto prihodek',
                         xaxis=dict(type='category'),
                         yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                    tickformat=',.0f', ticksuffix=' €'))
            st.plotly_chart(fig2, use_container_width=True)

    # TAB 2: Po virih / plačnikih
    with itab2:
        if 'vir' not in fi.columns:
            st.info("Ni stolpca 'Vir' v podatkih.")
        else:
            c1, c2 = st.columns(2)

            with c1:
                vir_data = fi.groupby('vir')['neto'].sum().reset_index().sort_values('neto', ascending=False)
                fig_vir = px.pie(vir_data, values='neto', names='vir',
                                 color_discrete_sequence=px.colors.qualitative.Set2,
                                 hole=0.35)
                fig_vir.update_traces(
                    textinfo='percent+label',
                    hovertemplate='<b>%{label}</b><br>%{value:,.0f} €  (%{percent})<extra></extra>'
                )
                chart_layout(fig_vir, title='Po viru prihodka', height=380)
                st.plotly_chart(fig_vir, use_container_width=True)

            with c2:
                if 'placnik' in fi.columns and fi['placnik'].notna().any():
                    pl_data = fi.groupby('placnik')['neto'].sum().reset_index().sort_values('neto', ascending=True)
                    fig_pl = go.Figure(go.Bar(
                        x=pl_data['neto'], y=pl_data['placnik'],
                        orientation='h', marker_color=C_TEAL, marker_opacity=0.85,
                        text=pl_data['neto'].apply(fmt_eur),
                        textposition='outside', textfont=dict(size=10),
                        hovertemplate='<b>%{y}</b><br>%{x:,.0f} €<extra></extra>'
                    ))
                    chart_layout(fig_pl, title='Po plačniku', height=380,
                                 xaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                            tickformat=',.0f', ticksuffix=' €'))
                    st.plotly_chart(fig_pl, use_container_width=True)
                else:
                    st.info("Ni podatkov o plačnikih.")

            # Tabela po viru
            vir_tbl = fi.groupby('vir')['neto'].agg(['sum', 'mean', 'count']).reset_index()
            vir_tbl.columns = ['Vir', 'Skupaj €', 'Povp. €', 'Mesecev']
            vir_tbl['Delež %'] = vir_tbl['Skupaj €'] / vir_tbl['Skupaj €'].sum() * 100
            st.dataframe(
                vir_tbl.style.format({'Skupaj €': '{:,.0f} €', 'Povp. €': '{:,.0f} €',
                                      'Delež %': '{:.1f}%'}),
                use_container_width=True, hide_index=True
            )

    # TAB 3: Raw tabela
    with itab3:
        disp_cols = [c for c in ['year_month', 'datum', 'vir', 'neto', 'placnik', 'komentar']
                     if c in fi.columns]
        rename = {'year_month': 'Mesec', 'datum': 'Datum', 'vir': 'Vir',
                  'neto': 'Neto €', 'placnik': 'Plačnik', 'komentar': 'Komentar'}
        st.dataframe(
            fi[disp_cols].sort_values('year_month', ascending=False)
                .rename(columns=rename)
                .style.format({'Neto €': '{:,.0f} €'}),
            use_container_width=True, hide_index=True, height=450
        )
        st.metric("Skupno vnosov", len(fi))


# ========== NET WORTH TRACKER ==========

def create_networth_tab(accounts_df):
    """Net Worth tracker po računih in tipu."""

    # Prazen state
    if accounts_df is None or accounts_df.empty:
        st.info("Še ni podatkov o računih.")
        st.markdown("""
        **Kako dodaš podatke:**
        1. Odpri Excel → sheet `accounts`
        2. Za vsak račun dodaj vrstico:

        | ime_racuna | tip | valuta | datum_stanja | stanje | aktiven | opombe |
        |---|---|---|---|---|---|---|
        | NLB osebni | banka | EUR | 2026-04-15 | 5250.00 | 1 | |
        | IBKR | investicije | EUR | 2026-04-15 | 18400.00 | 1 | |
        | ZPIZ | pokojnina | EUR | 2026-04-15 | 3200.00 | 1 | |

        3. Shrani Excel in **Uvozi v bazo** (sidebar)

        **Tipi računov** (priporočeni): `banka`, `investicije`, `pokojnina`, `nepremičnine`, `gotovina`
        """)
        return

    acc = accounts_df.copy()

    # Normalizacija stolpcev
    acc.columns = [c.strip().lower().replace(' ', '_') for c in acc.columns]

    # Pretvori datum in stanje
    if 'datum_stanja' in acc.columns:
        acc['datum_stanja'] = pd.to_datetime(acc['datum_stanja'], errors='coerce')
    if 'stanje' in acc.columns:
        acc['stanje'] = pd.to_numeric(acc['stanje'], errors='coerce').fillna(0)
    if 'aktiven' in acc.columns:
        acc['aktiven'] = acc['aktiven'].astype(str).str.strip().isin(['1', 'True', 'true', 'DA', 'da', 'yes'])
    else:
        acc['aktiven'] = True

    aktivni = acc[acc['aktiven']].copy()

    if aktivni.empty:
        st.warning("Vsi računi so označeni kot neaktivni.")
        return

    # ---- ZADNJE STANJE PO RAČUNU ----
    # Če je več vnosov za isti račun, vzami najnovejšega
    if 'datum_stanja' in acc.columns:
        latest = (aktivni.sort_values('datum_stanja', ascending=False)
                  .groupby('ime_racuna', as_index=False).first())
    else:
        latest = aktivni.groupby('ime_racuna', as_index=False).last()

    total_nw = latest['stanje'].sum()

    # ---- METRIKE ----
    tip_groups = latest.groupby('tip')['stanje'].sum() if 'tip' in latest.columns else pd.Series()

    # Dinamične metrike po tipu
    tip_cols = st.columns(min(len(tip_groups) + 1, 5))
    with tip_cols[0]:
        st.metric("💎 Skupni Net Worth", f"{total_nw:,.0f} €")
    for i, (tip, val) in enumerate(tip_groups.items(), 1):
        if i < len(tip_cols):
            icons = {'banka': '🏦', 'investicije': '📈', 'pokojnina': '🏛️',
                     'nepremičnine': '🏠', 'gotovina': '💵'}
            icon = icons.get(str(tip).lower(), '💼')
            with tip_cols[i]:
                st.metric(f"{icon} {str(tip).capitalize()}", f"{val:,.0f} €")

    st.divider()

    # ---- GRAFI ----
    nw_tab1, nw_tab2, nw_tab3 = st.tabs(["📊 Pregled", "📈 Trend čez čas", "📋 Računi"])

    # TAB: Pregled
    with nw_tab1:
        c1, c2 = st.columns(2)

        with c1:
            if 'tip' in latest.columns:
                fig_pie = px.pie(
                    latest, values='stanje', names='tip',
                    color_discrete_sequence=px.colors.qualitative.Set2,
                    hole=0.4
                )
                fig_pie.update_traces(
                    textinfo='percent+label',
                    hovertemplate='<b>%{label}</b><br>%{value:,.0f} €  (%{percent})<extra></extra>'
                )
                chart_layout(fig_pie, title='Razporeditev po tipu', height=380)
                st.plotly_chart(fig_pie, use_container_width=True)

        with c2:
            latest_sorted = latest.sort_values('stanje', ascending=True)
            tip_color_map = {}
            for t in (latest_sorted['tip'].unique() if 'tip' in latest_sorted.columns else []):
                if t not in tip_color_map:
                    tip_color_map[t] = PALETTE[len(tip_color_map) % len(PALETTE)]
            bar_colors = ([tip_color_map.get(t, C_BLUE) for t in latest_sorted['tip']]
                          if 'tip' in latest_sorted.columns else C_BLUE)

            fig_bar = go.Figure(go.Bar(
                x=latest_sorted['stanje'], y=latest_sorted['ime_racuna'],
                orientation='h', marker_color=bar_colors, marker_opacity=0.85,
                text=latest_sorted['stanje'].apply(fmt_eur),
                textposition='outside', textfont=dict(size=10),
                hovertemplate='<b>%{y}</b><br>%{x:,.0f} €<extra></extra>'
            ))
            chart_layout(fig_bar, title='Stanje po računih', height=380,
                         margin=dict(l=8, r=70, t=48, b=30),
                         xaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                    tickformat=',.0f', ticksuffix=' €'))
            st.plotly_chart(fig_bar, use_container_width=True)

    # TAB: Trend čez čas
    with nw_tab2:
        if 'datum_stanja' not in acc.columns or acc['datum_stanja'].isna().all():
            st.info("Za trend potrebuješ več vnosov z različnimi datumi.")
        else:
            trend = (aktivni.groupby('datum_stanja')['stanje'].sum()
                     .reset_index().sort_values('datum_stanja'))

            if len(trend) < 2:
                st.info("Za trend grafikona potrebuješ vsaj 2 datuma z vnosi.")
                st.dataframe(trend.rename(columns={'datum_stanja': 'Datum', 'stanje': 'Net Worth €'}),
                             use_container_width=True, hide_index=True)
            else:
                first_val = trend['stanje'].iloc[0]
                last_val  = trend['stanje'].iloc[-1]
                change     = last_val - first_val
                change_pct = (change / first_val * 100) if first_val > 0 else 0

                fig_trend = go.Figure()
                fig_trend.add_trace(go.Scatter(
                    x=trend['datum_stanja'], y=trend['stanje'],
                    mode='lines+markers', name='Net Worth',
                    line=dict(color=C_BLUE, width=2.5),
                    marker=dict(size=7, color=C_BLUE),
                    fill='tozeroy', fillcolor='rgba(31,78,120,0.08)',
                    hovertemplate='<b>%{x|%d.%m.%Y}</b><br>%{y:,.0f} €<extra></extra>'
                ))
                chart_layout(fig_trend,
                             title=f'Net Worth trend — sprememba: {change:+,.0f} € ({change_pct:+.1f}%)',
                             hovermode='x unified',
                             yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                        tickformat=',.0f', ticksuffix=' €'))
                st.plotly_chart(fig_trend, use_container_width=True)

                # Trend po tipu računov
                if 'tip' in aktivni.columns:
                    st.subheader("Trend po tipu računa")
                    tip_trend = (aktivni.groupby(['datum_stanja', 'tip'])['stanje']
                                 .sum().reset_index().sort_values('datum_stanja'))

                    fig_tip = px.line(
                        tip_trend, x='datum_stanja', y='stanje', color='tip',
                        markers=True,
                        labels={'datum_stanja': '', 'stanje': '', 'tip': 'Tip'},
                        color_discrete_sequence=px.colors.qualitative.Set2
                    )
                    fig_tip.update_traces(
                        hovertemplate='<b>%{x|%d.%m.%Y}</b><br>%{y:,.0f} €<extra></extra>',
                        line=dict(width=2.5), marker=dict(size=6)
                    )
                    chart_layout(fig_tip, title='Trend po tipu računa',
                                 yaxis=dict(showgrid=True, gridcolor='rgba(128,128,128,0.15)',
                                            tickformat=',.0f', ticksuffix=' €'))
                    st.plotly_chart(fig_tip, use_container_width=True)

    # TAB: Tabela računov
    with nw_tab3:
        st.subheader("📋 Vsi aktivni računi")
        show_cols = [c for c in ['ime_racuna', 'tip', 'valuta', 'datum_stanja', 'stanje', 'opombe']
                     if c in latest.columns]
        rename = {'ime_racuna': 'Račun', 'tip': 'Tip', 'valuta': 'Valuta',
                  'datum_stanja': 'Datum stanja', 'stanje': 'Stanje €', 'opombe': 'Opombe'}
        display = latest[show_cols].rename(columns=rename).sort_values('Stanje €', ascending=False)

        fmt = {'Stanje €': '{:,.2f} €'}
        if 'Datum stanja' in display.columns:
            display['Datum stanja'] = display['Datum stanja'].dt.strftime('%d.%m.%Y')

        st.dataframe(display.style.format(fmt), use_container_width=True, hide_index=True)

        # Skupaj po tipu
        if 'Tip' in display.columns:
            st.subheader("Skupaj po tipu")
            by_tip = display.groupby('Tip')['Stanje €'].sum().reset_index()
            by_tip['Delež %'] = by_tip['Stanje €'] / by_tip['Stanje €'].sum() * 100
            st.dataframe(
                by_tip.style.format({'Stanje €': '{:,.2f} €', 'Delež %': '{:.1f}%'}),
                use_container_width=True, hide_index=True
            )

        neaktivni = acc[~acc['aktiven']]
        if not neaktivni.empty:
            with st.expander(f"Neaktivni računi ({len(neaktivni)})", expanded=False):
                st.dataframe(neaktivni[show_cols].rename(columns=rename),
                             use_container_width=True, hide_index=True)


# ========== EXCEL EXPORT ==========

def generate_excel_dashboard(df, budget_df=None):
    """Generiraj Excel dashboard."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Dashboard')
    ws['B2'] = 'FINANČNI DASHBOARD'
    ws['B2'].font = Font(bold=True, size=16, color='1F4E78')
    ws.merge_cells('B2:E2')
    ws['B3'] = f'Generirano: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['B3'].font = Font(italic=True)

    prihodki_sum = df[df['tip'].str.lower() == 'prihodek']['znesek'].sum()
    odhodki_sum = df[df['tip'].str.lower() == 'odhodek']['znesek'].sum()
    bilanca = prihodki_sum - odhodki_sum

    row = 5
    for label, val in [('Skupaj prihodki:', prihodki_sum),
                        ('Skupaj odhodki:', odhodki_sum),
                        ('Bilanca:', bilanca)]:
        ws[f'B{row}'] = label
        ws[f'C{row}'] = val
        ws[f'C{row}'].number_format = '#,##0.00 €'
        if label == 'Bilanca:':
            ws[f'C{row}'].font = Font(bold=True,
                                       color='548235' if bilanca >= 0 else 'C00000')
        row += 1

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ========== MAIN APP ==========

st.markdown('<div class="main-header">💶 Osebni finančni dashboard</div>', unsafe_allow_html=True)
last_imp = get_last_import()
if last_imp:
    st.markdown(f'<div class="main-subheader">Podatki: {last_imp}</div>', unsafe_allow_html=True)

# ---- SIDEBAR ----
with st.sidebar:
    st.header("🗄️ Baza podatkov")

    if db_exists():
        last_import = get_last_import()
        db_size = get_db_size_kb()
        st.markdown(
            f'<div class="db-status-ok">✅ Baza aktivna &nbsp;·&nbsp; {db_size} KB<br>'
            f'<small>Zadnji uvoz: {last_import or "neznano"}</small></div>',
            unsafe_allow_html=True
        )
        # Sheet status
        ss = st.session_state
        sheet_info = [
            ("💳 Transakcije",  ss.df,         lambda d: f"{len(d):,} vrstic"),
            ("🏢 S.P.",         ss.sp_df,       lambda d: f"{len(d):,} vrstic"),
            ("📋 Budget plan",  ss.budget_df,   lambda d: f"{len(d):,} kategorij"),
            ("💰 Prihodki",     ss.income_df,   lambda d: f"{len(d):,} vrstic"),
            ("🏦 Računi",       ss.accounts_df, lambda d: f"{len(d):,} vrstic"),
        ]
        lines = []
        for label, data, fmt in sheet_info:
            if data is not None and not data.empty:
                lines.append(f"✅ {label} &nbsp;<small>({fmt(data)})</small>")
            else:
                lines.append(f"⬜ {label}")
        st.markdown(
            '<div class="sheet-status">' + "<br>".join(lines) + '</div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div class="db-status-empty">⚠️ Baza prazna<br>'
            '<small>Uvozi Excel datoteko</small></div>',
            unsafe_allow_html=True
        )

    st.divider()
    st.header("📁 Uvoz Excel datoteke")
    uploaded_file = st.file_uploader(
        "Naloži Excel (vseh 5 sheetov)", type=['xlsx', 'xls'],
        help="Excel mora vsebovati sheete: transactions, sp_transactions, budget_plan, income_history, accounts"
    )

    if uploaded_file:
        if st.button("🔄 Uvozi in shrani v bazo", type="primary"):
            with st.spinner("Uvažam podatke..."):
                raw_data, missing = load_excel_multisheet(uploaded_file)
                if raw_data:
                    ok, err = save_to_db(raw_data)
                    if ok:
                        apply_prepared_data(raw_data)
                        if missing:
                            st.warning(f"Manjkajoči sheeti: {', '.join(missing)}")
                        else:
                            st.success("✅ Vsi podatki uvoženi in shranjeni!")
                        st.rerun()
                    else:
                        st.error(f"Napaka pri shranjevanju: {err}")

    if db_exists():
        st.divider()
        if st.button("🗑️ Počisti bazo", help="Izbriše vse podatke iz baze"):
            clear_db()
            for key in ['df', 'sp_df', 'budget_df', 'income_df', 'accounts_df', 'raw_data']:
                st.session_state[key] = None if key != 'raw_data' else {}
            st.session_state.db_loaded = False
            st.rerun()

    st.divider()

    # Filtri (samo če so transactions naložene)
    if st.session_state.df is not None:
        st.header("🎛️ Filtri")

        df = st.session_state.df

        cur_year  = datetime.now().year
        cur_month = datetime.now().month

        available_years = sorted(df['leto'].unique().tolist()) if 'leto' in df.columns else []
        default_years   = [cur_year] if cur_year in available_years else available_years[-1:]
        selected_years  = st.multiselect("Izberi leta", available_years, default=default_years)

        available_months = sorted(df['mesec'].unique().tolist()) if 'mesec' in df.columns else []
        month_names = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Maj', 6: 'Jun',
                       7: 'Jul', 8: 'Avg', 9: 'Sep', 10: 'Okt', 11: 'Nov', 12: 'Dec'}
        # Default: YTD (jan–trenutni mesec), samo če je izbrano samo tekoče leto
        default_months = (list(range(1, cur_month + 1))
                          if selected_years == [cur_year] else [])
        selected_months = st.multiselect(
            "Izberi mesece", available_months,
            default=default_months,
            format_func=lambda x: f"{x} - {month_names.get(x, '')}"
        )

        available_categories = (sorted(df['kategorija'].dropna().unique().tolist())
                                 if 'kategorija' in df.columns else [])
        selected_categories = st.multiselect("Izberi kategorije", available_categories,
                                             help="Pusti prazno za vse kategorije")

        if selected_categories:
            filtered_for_subcat = df[df['kategorija'].isin(selected_categories)]
            available_subcategories = (
                sorted(filtered_for_subcat['podkategorija'].dropna().unique().tolist())
                if 'podkategorija' in filtered_for_subcat.columns else []
            )
        else:
            available_subcategories = (
                sorted(df['podkategorija'].dropna().unique().tolist())
                if 'podkategorija' in df.columns else []
            )
        selected_subcategories = st.multiselect("Izberi podkategorije", available_subcategories,
                                                help="Pusti prazno za vse podkategorije")

        st.divider()
        st.header("📊 Prikaz")
        view_type = st.radio("Tip prikaza", ['Trend', 'Vsota'], horizontal=True)
        view_type = 'trend' if view_type == 'Trend' else 'sum'

        chart_type = st.radio("Tip grafa kategorij", ['Pie', 'Bar'], horizontal=True).lower()

        group_by_year = False
        if len(selected_years) > 1:
            group_by_year = st.checkbox(
                "Prikaži kategorije po letih", value=True,
                help="Primerja vsako kategorijo po izbranih letih"
            )
    else:
        selected_years = []
        selected_months = []
        selected_categories = []
        selected_subcategories = []
        view_type = 'trend'
        chart_type = 'pie'
        group_by_year = False


# ---- MAIN CONTENT ----

if st.session_state.df is None:
    st.markdown("---")
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.markdown("### 👈 Začni z uvozom Excel datoteke")
        st.markdown("""
        **Korak 1** — Pripravi Excel z naslednjimi sheeti:

        | Sheet | Vsebina |
        |---|---|
        | `transactions` | Osebne finance |
        | `sp_transactions` | S.P. prihodki in odhodki |
        | `budget_plan` | Letni plan po kategorijah |
        | `income_history` | Plača / prihodki tracking |
        | `accounts` | Računi za Net Worth |

        **Korak 2** — Naloži datoteko v stransko vrstico

        **Korak 3** — Pritisni **Uvozi in shrani v bazo**

        ✅ Ob naslednjem zagonu se podatki naložijo **avtomatsko**.
        """)

else:
    df = st.session_state.df

    # Filtrirani podatki za metrike
    filtered = df.copy()
    if selected_years:
        filtered = filtered[filtered['leto'].isin(selected_years)]
    if selected_months:
        filtered = filtered[filtered['mesec'].isin(selected_months)]
    if selected_categories:
        filtered = filtered[filtered['kategorija'].isin(selected_categories)]
    if selected_subcategories:
        filtered = filtered[filtered['podkategorija'].isin(selected_subcategories)]

    st_mes = filtered['YearMonth_sort'].nunique() if 'YearMonth_sort' in filtered.columns else 0

    # ---- TEKOČA LETNA BILANCA ----
    bil_col, chk_col = st.columns([3, 1])
    with bil_col:
        st.subheader("Tekoča letna bilanca")
    with chk_col:
        vkljuci_sp = st.checkbox("+ S.P.", value=False,
                                 help="Vključi prihodke in odhodke S.P. v bilanco")

    sorted_years = sorted(selected_years) if selected_years else []

    def calc_bilanca(year, months):
        """Izračuna prihodke, odhodke, bilanco za leto/mesece, opcijsko + S.P."""
        yd = filtered[filtered['leto'] == year]
        if months:
            yd = yd[yd['mesec'].isin(months)]
        yp = yd[yd['tip'].str.lower() == 'prihodek']['znesek'].sum()
        yo = yd[yd['tip'].str.lower() == 'odhodek']['znesek'].sum()

        if vkljuci_sp and st.session_state.sp_df is not None:
            sp = st.session_state.sp_df
            sp_y = sp[sp['leto'] == year]
            if months:
                sp_y = sp_y[sp_y['mesec'].isin(months)]
            yp += sp_y[sp_y['tip'].str.lower() == 'prihodek']['znesek'].sum()
            yo += sp_y[sp_y['tip'].str.lower() == 'odhodek']['znesek'].sum()

        return yp, yo, yp - yo

    if len(sorted_years) == 1:
        year = sorted_years[0]
        yp, yo, yb = calc_bilanca(year, selected_months)
        ca, cb, cc, cd, _ = st.columns([1, 1, 1, 1, 1])
        with ca:
            st.metric("💰 Prihodki", f"{yp:,.0f} €")
        with cb:
            st.metric("💸 Odhodki", f"{yo:,.0f} €")
        with cc:
            st.metric("📊 Bilanca", f"{yb:,.0f} €",
                      delta="✅ Presežek" if yb >= 0 else "⚠️ Primanjkljaj",
                      delta_color="normal" if yb >= 0 else "inverse")
        with cd:
            st.metric("📅 Mesecev", st_mes)

    elif len(sorted_years) <= 4:
        cols_letna = st.columns(len(sorted_years))
        for idx, year in enumerate(sorted_years):
            yp, yo, yb = calc_bilanca(year, selected_months)
            with cols_letna[idx]:
                st.markdown(f"### {year}")
                st.metric("Prihodki", f"{yp:,.0f} €")
                st.metric("Odhodki", f"{yo:,.0f} €")
                st.metric("Bilanca", f"{yb:,.0f} €",
                          delta="✅ Presežek" if yb >= 0 else "⚠️ Primanjkljaj",
                          delta_color="normal" if yb >= 0 else "inverse")
    else:
        rows = []
        for year in sorted_years:
            yp, yo, yb = calc_bilanca(year, selected_months)
            rows.append({'Leto': year, 'Prihodki €': yp, 'Odhodki €': yo, 'Bilanca €': yb})
        annual_tbl = pd.DataFrame(rows)

        def color_annual(row):
            if row['Bilanca €'] < 0:
                return ['', '', '', 'color: #ff6b6b; font-weight: 600']
            return ['', '', '', 'color: #6abf69; font-weight: 600']

        st.dataframe(
            annual_tbl.style.apply(color_annual, axis=1)
                .format({'Prihodki €': '{:,.0f} €', 'Odhodki €': '{:,.0f} €',
                         'Bilanca €': '{:,.0f} €'}),
            use_container_width=True, hide_index=True
        )

    st.divider()

    # ---- TABI ----
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📈 Mesečni trend",
        "🏷️ Kategorije",
        "📋 Budget",
        "🏢 S.P. analiza",
        "💰 Prihodki",
        "💎 Net Worth",
        "📋 Podatki",
        "🗄️ Baza podatkov"
    ])

    # TAB 1: Mesečni trend
    with tab1:
        fig = create_monthly_trend(df, selected_years, selected_months, view_type)
        st.plotly_chart(fig, use_container_width=True)

        if view_type == 'trend':
            st.subheader("📊 Mesečna tabela")
            filtered_trend = df.copy()
            if selected_years:
                filtered_trend = filtered_trend[filtered_trend['leto'].isin(selected_years)]
            if selected_months:
                filtered_trend = filtered_trend[filtered_trend['mesec'].isin(selected_months)]

            prihodki_mes = (filtered_trend[filtered_trend['tip'].str.lower() == 'prihodek']
                            .groupby('YearMonth_sort')['znesek'].sum())
            odhodki_mes = (filtered_trend[filtered_trend['tip'].str.lower() == 'odhodek']
                           .groupby('YearMonth_sort')['znesek'].sum())

            trend_table = pd.DataFrame({
                'Prihodki': prihodki_mes,
                'Odhodki': odhodki_mes,
            }).fillna(0).sort_index()
            trend_table['Bilanca'] = trend_table['Prihodki'] - trend_table['Odhodki']

            total_row = pd.DataFrame({
                'Prihodki': [trend_table['Prihodki'].sum()],
                'Odhodki': [trend_table['Odhodki'].sum()],
                'Bilanca': [trend_table['Bilanca'].sum()]
            }, index=['═══ SKUPAJ ═══'])

            trend_table = pd.concat([trend_table, total_row])

            def highlight_total(row):
                if row.name == '═══ SKUPAJ ═══':
                    return ['background-color: #1F4E78; color: white; font-weight: bold'] * len(row)
                return [''] * len(row)

            st.dataframe(
                trend_table.style.format("{:.2f} €").apply(highlight_total, axis=1),
                use_container_width=True
            )

    # TAB 2: Kategorije
    with tab2:
        fig_cat = create_category_chart(
            df, selected_years, selected_months,
            selected_categories, chart_type, group_by_year
        )
        if fig_cat:
            st.plotly_chart(fig_cat, use_container_width=True)

            st.subheader("📊 TOP 15 kategorij")
            filtered_cat = df[df['tip'].str.lower() == 'odhodek'].copy()
            if selected_years:
                filtered_cat = filtered_cat[filtered_cat['leto'].isin(selected_years)]
            if selected_months:
                filtered_cat = filtered_cat[filtered_cat['mesec'].isin(selected_months)]
            if selected_categories:
                filtered_cat = filtered_cat[filtered_cat['kategorija'].isin(selected_categories)]
            if selected_subcategories:
                filtered_cat = filtered_cat[filtered_cat['podkategorija'].isin(selected_subcategories)]

            if 'kategorija' in filtered_cat.columns:
                cat_summary = filtered_cat.groupby('kategorija').agg(
                    {'znesek': ['sum', 'count', 'mean']}
                ).round(2)
                cat_summary.columns = ['Skupaj €', 'Število', 'Povprečje €']
                cat_summary = cat_summary.sort_values('Skupaj €', ascending=False).head(15)

                total_row = pd.DataFrame({
                    'Skupaj €': [cat_summary['Skupaj €'].sum()],
                    'Število': [cat_summary['Število'].sum()],
                    'Povprečje €': [cat_summary['Skupaj €'].sum() / cat_summary['Število'].sum()]
                }, index=['═══ SKUPAJ (TOP 15) ═══'])

                cat_summary = pd.concat([cat_summary, total_row])

                def highlight_total_cat(row):
                    if row.name == '═══ SKUPAJ (TOP 15) ═══':
                        return ['background-color: #1F4E78; color: white; font-weight: bold'] * len(row)
                    return [''] * len(row)

                st.dataframe(
                    cat_summary.style.apply(highlight_total_cat, axis=1),
                    use_container_width=True
                )

    # TAB 3: Budget tracking
    with tab3:
        create_budget_tab(df, st.session_state.budget_df)

    # TAB 4: S.P. analiza
    with tab4:
        create_sp_tab(st.session_state.sp_df)

    # TAB 5: Prihodki
    with tab5:
        create_income_tab(st.session_state.income_df)

    # TAB 6: Net Worth
    with tab6:
        create_networth_tab(st.session_state.accounts_df)

    # TAB 7: Raw podatki
    with tab7:
        st.subheader("📋 Raw podatki")
        display_cols = ['YearMonth_sort', 'datum', 'tip', 'znesek',
                        'kategorija', 'podkategorija', 'zadeva', 'nujnost']
        display_cols = [col for col in display_cols if col in filtered.columns]

        st.dataframe(
            filtered[display_cols].sort_values('YearMonth_sort', ascending=False),
            use_container_width=True, height=500
        )
        st.metric("Skupno transakcij", len(filtered))

    # TAB 8: Baza podatkov
    with tab8:
        st.subheader("🗄️ Status baze podatkov")

        db_info = get_db_info()
        last_import = get_last_import()
        db_size = get_db_size_kb()

        col_meta1, col_meta2 = st.columns(2)
        with col_meta1:
            st.metric("Zadnji uvoz", last_import or "—")
        with col_meta2:
            st.metric("Velikost baze", f"{db_size} KB")

        st.divider()
        st.subheader("📋 Tabele v bazi")

        table_labels = {
            'transactions': '💳 Transakcije (osebne finance)',
            'sp_transactions': '🏢 S.P. transakcije',
            'budget_plan': '📋 Plan po kategorijah',
            'income_history': '💰 Plača/Prihodki',
            'accounts': '🏦 Računi (Net Worth)',
        }

        for table_key in EXPECTED_TABLES:
            label = table_labels.get(table_key, table_key)
            if table_key in db_info:
                info = db_info[table_key]
                with st.expander(f"✅ {label} — {info['rows']} vrstic, {info['cols']} stolpcev"):
                    st.markdown(f"**Stolpci:** {', '.join(f'`{c}`' for c in info['columns'])}")

                    # Prikaži vzorec podatkov
                    raw = st.session_state.raw_data.get(table_key)
                    if raw is not None and not raw.empty:
                        st.dataframe(raw.head(5), use_container_width=True)
            else:
                st.markdown(f"⬜ {label} — *ni podatkov*")

    # Download Excel
    st.divider()
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("📥 Generiraj in prenesi Excel dashboard",
                     type="primary", use_container_width=True):
            with st.spinner("Generiram dashboard..."):
                excel_file = generate_excel_dashboard(df, st.session_state.budget_df)
                timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")

                st.download_button(
                    label="💾 Prenesi Dashboard",
                    data=excel_file,
                    file_name=f"Dashboard_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
